"""
document_loader.py
Scans the documents/ folder, loads every supported file, splits the text
into chunks and returns a list of LangChain Document objects.

Supported formats:
  - PDF  (.pdf)
  - Word (.docx)
  - Plain text (.txt, .md, .csv)
  - Excel (.xls, .xlsx)
"""
import logging
import os
from pathlib import Path
from typing import Any, List

from langchain_community.document_loaders import Docx2txtLoader, PyPDFLoader, TextLoader
from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter

from config import settings

logger = logging.getLogger(__name__)


class ExcelLoader:
    """Load Excel workbooks into one LangChain Document per sheet."""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)

    def load(self) -> List[Document]:
        extension = self.file_path.suffix.lower()
        if extension == ".xlsx":
            return self._load_xlsx()
        if extension == ".xls":
            return self._load_xls()
        raise ValueError(f"Unsupported Excel file type: {extension}")

    def _load_xlsx(self) -> List[Document]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to read .xlsx files.") from exc

        workbook = load_workbook(
            filename=str(self.file_path),
            read_only=True,
            data_only=True,
        )

        docs: List[Document] = []
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                rows = [
                    self._row_to_text(row)
                    for row in worksheet.iter_rows(values_only=True)
                ]
                docs.append(self._make_sheet_document(sheet_name, rows))
        finally:
            workbook.close()

        return docs

    def _load_xls(self) -> List[Document]:
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError("xlrd is required to read .xls files.") from exc

        workbook = xlrd.open_workbook(str(self.file_path))
        docs: List[Document] = []

        for sheet_name in workbook.sheet_names():
            worksheet = workbook.sheet_by_name(sheet_name)
            rows = [
                self._row_to_text(worksheet.row_values(row_idx))
                for row_idx in range(worksheet.nrows)
            ]
            docs.append(self._make_sheet_document(sheet_name, rows))

        return docs

    def _make_sheet_document(self, sheet_name: str, rows: List[str]) -> Document:
        non_empty_rows = [row for row in rows if row]
        content = "\n".join(non_empty_rows).strip()
        if not content:
            content = f"[Empty sheet: {sheet_name}]"

        return Document(
            page_content=content,
            metadata={
                "sheet_name": sheet_name,
                "source": str(self.file_path),
                "filename": self.file_path.name,
                "filetype": self.file_path.suffix.lstrip("."),
            },
        )

    @staticmethod
    def _row_to_text(row: Any) -> str:
        values = [str(cell).strip() for cell in row if cell not in (None, "")]
        return " | ".join(values)


LOADER_MAP = {
    ".pdf": PyPDFLoader,
    ".docx": Docx2txtLoader,
    ".txt": TextLoader,
    ".md": TextLoader,
    ".csv": TextLoader,
    ".xls": ExcelLoader,
    ".xlsx": ExcelLoader,
}
SUPPORTED_EXTENSIONS = set(LOADER_MAP.keys())


def _sanitise(text: str) -> str:
    """
    Remove Unicode surrogate characters and any other code points that cannot
    be encoded as UTF-8. These sometimes appear when PDFs contain corrupt or
    scanned text.
    """
    return text.encode("utf-8", errors="replace").decode("utf-8", errors="replace")


def load_documents_from_folder(folder: str | None = None) -> List[Document]:
    """
    Walk *folder* (defaults to settings.DOCUMENTS_DIR), load every
    supported file and return the raw Document list (not yet split).
    """
    folder = folder or settings.DOCUMENTS_DIR
    docs: List[Document] = []

    if not os.path.isdir(folder):
        logger.warning("Documents folder '%s' does not exist - creating it.", folder)
        os.makedirs(folder, exist_ok=True)
        return docs

    for root, _, files in os.walk(folder):
        for filename in files:
            filepath = Path(root) / filename
            ext = filepath.suffix.lower()
            loader_cls = LOADER_MAP.get(ext)
            if loader_cls is None:
                logger.debug("Skipping unsupported file: %s", filepath)
                continue
            try:
                logger.info("Loading: %s", filepath)
                loader = loader_cls(str(filepath))
                loaded = loader.load()
                for doc in loaded:
                    doc.page_content = _sanitise(doc.page_content)
                    doc.metadata.update(
                        {
                            "source": str(filepath),
                            "filename": filename,
                            "filetype": ext.lstrip("."),
                        }
                    )
                docs.extend(loaded)
            except Exception as exc:
                logger.error("Failed to load '%s': %s", filepath, exc)

    logger.info("Loaded %d raw document page(s) from '%s'.", len(docs), folder)
    return docs


def split_documents(docs: List[Document]) -> List[Document]:
    """Split documents into chunks suitable for embedding."""
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=settings.CHUNK_SIZE,
        chunk_overlap=settings.CHUNK_OVERLAP,
        separators=["\n\n", "\n", " ", ""],
    )
    chunks = splitter.split_documents(docs)
    logger.info("Split into %d chunk(s).", len(chunks))
    return chunks


def load_and_split(folder: str | None = None) -> List[Document]:
    """Convenience function: load + split in one call."""
    docs = load_documents_from_folder(folder)
    if not docs:
        return []
    return split_documents(docs)
